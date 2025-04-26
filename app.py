from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageOps, ImageEnhance
import os

app = Flask(__name__)
app.secret_key = 'a02e0001d7cb2e08646b5dcd27d10d6e'

# Paths
EXCEL_FILE = "users.xlsx"
UPLOAD_FOLDER = 'static/uploads'

# Create folders/files if not exist
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Password"])
    wb.save(EXCEL_FILE)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Function to apply vintage filter
def apply_vintage_filter(image_path):
    img = Image.open(image_path)
    img = img.convert('L')  # Grayscale
    img = ImageOps.colorize(img, black="black", white="#d6a15c")  # Warm tint
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.5)
    filtered_image_path = os.path.join(UPLOAD_FOLDER, 'vintage_' + os.path.basename(image_path))
    img.save(filtered_image_path)
    return filtered_image_path

# Routes

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Check if username exists
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                flash("Username already exists!")
                return redirect(url_for('signup'))

        ws.append([username, password])
        wb.save(EXCEL_FILE)

        flash("Account created successfully! Please login.")
        return redirect(url_for('login'))

    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                flash(f"Welcome, {username}!")
                return redirect(url_for('vintage_home'))

        flash("Invalid username or password.")
        return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/vintage')
def vintage_home():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(url_for('vintage_home'))

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(url_for('vintage_home'))

    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)

        filtered_path = apply_vintage_filter(filepath)

        # Convert path to relative for rendering
        filtered_rel_path = '/' + filtered_path.replace('\\', '/')
        return render_template('index.html', filtered_image_url=filtered_rel_path)

@app.route('/static/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

if __name__ == '__main__':
    app.run(debug=True)
